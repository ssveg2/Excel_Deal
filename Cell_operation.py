import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active                                              #获取当前活动工作表，不需要括号
sheet.cell(row=1, column=1).value = 200                        #给单元格A1赋值200
sheet['A2'].value = 300                                        #给单元格A2赋值300
sheet['A3'].value = '=SUM(A1:A2)'                              #给单元格赋值等式
sheet.row_dimensions[1].height = 70                            #设置第一行行高为70
sheet.column_dimensions['B'].width = 20                        #设置B列宽度为20
sheet.merge_cells('A1:D1')                                     #合并单元格
sheet.unmerge_cells('A1:D1')                                   #拆分单元格
sheet.freeze_panes = 'A2'                                      #冻结第一行
sheet.freeze_panes = 'B1'                                      #冻结第一列
wb.save('sum.xlsx')                                            #保存Excel文件
wbFormulas = openpyxl.load_workbook('sum.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)
wbData = openpyxl.load_workbook('sum.xlsx', data_only=True)
sheet = wbData.active
print(sheet['A3'].value)



