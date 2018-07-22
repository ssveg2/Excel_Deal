import openpyxl
wb = openpyxl.Workbook()                        #新建Excel文件
wb.create_sheet(index=2, title='PS')            #创建工作表，参数为位置和名称
wb.create_sheet(index=1, title='SH')            #第二个工作表
sheet = wb.worksheets[0]                        #获取第一个工作表
sheet.title = 'PATAC'                           #给工作表命名
del wb['SH']                                    #删除指定名称工作表
sheet['A1'] = 'Hello PATAC!'                    #为指定单元格输入值
wb.save('example_write.xlsx')                   #保存Excel文件