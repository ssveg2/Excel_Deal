import openpyxl
wb = openpyxl.load_workbook('G:/Python_Learning/TEC_Python/automate_online-materials/produceSales.xlsx')
sheet = wb.active                                                                 #获取当前活动工作表对象
Price_Updates = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}                                                   #定义新的价格字典
for rowNum in range(2, sheet.max_row + 1):                                        #遍历活动工作表所有行
    #produceName = sheet.cell(row=rowNum, column=1).value                         #获取单元格的值
    produceName = sheet['A' + str(rowNum)].value
    if produceName in Price_Updates:                                              #识别名称是否在字典中
        sheet.cell(row=rowNum, column=2).value = Price_Updates[produceName]       #将键值赋值给需要的位置
wb.save('produceSalesUpdate2.xlsx')


